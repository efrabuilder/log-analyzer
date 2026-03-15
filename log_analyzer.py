"""
log_analyzer.py — Efraín Rojas Artavia
CLI tool that parses network device logs, detects anomalies,
and sends email alerts when thresholds are exceeded.
"""

import re
import smtplib
import logging
import argparse
from datetime import datetime
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import CONFIG

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("analyzer.log"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Log Patterns ───────────────────────────────────────────────────────────────
PATTERNS = {
    "interface_down": re.compile(
        r"(?P<timestamp>\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2}).*"
        r"(?P<device>\S+).*"
        r"(?:interface|line protocol|Interface)\s+(?P<interface>\S+).*"
        r"(?:down|DOWN|went down)",
        re.IGNORECASE
    ),
    "high_cpu": re.compile(
        r"(?P<timestamp>\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2}).*"
        r"CPU\s+(?:utilization|usage)[:\s]+(?P<value>\d+)%",
        re.IGNORECASE
    ),
    "authentication_failure": re.compile(
        r"(?P<timestamp>\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2}).*"
        r"(?:authentication failure|login failed|invalid password|auth fail)"
        r".*(?:user|from)?\s*(?P<user>\S+)?",
        re.IGNORECASE
    ),
    "link_flap": re.compile(
        r"(?P<timestamp>\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2}).*"
        r"(?P<interface>\S+).*(?:changed state|flap|up/down)",
        re.IGNORECASE
    ),
    "memory_warning": re.compile(
        r"(?P<timestamp>\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2}).*"
        r"(?:memory|mem)\s+(?:warning|critical|low|usage)[:\s]+(?P<value>\d+)%",
        re.IGNORECASE
    ),
    "ospf_neighbor_down": re.compile(
        r"(?P<timestamp>\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2}).*"
        r"OSPF.*neighbor.*(?:down|dead|timeout)",
        re.IGNORECASE
    ),
    "bgp_session_drop": re.compile(
        r"(?P<timestamp>\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2}).*"
        r"BGP.*(?:session|peer).*(?:dropped|down|reset|closed)",
        re.IGNORECASE
    ),
    "error_generic": re.compile(
        r"(?P<timestamp>\w{3}\s+\d+\s+\d{2}:\d{2}:\d{2}).*"
        r"\b(?:ERROR|CRITICAL|FATAL|EMERG|ALERT)\b",
        re.IGNORECASE
    ),
}

SEVERITY = {
    "interface_down":       "CRITICAL",
    "high_cpu":             "WARNING",
    "authentication_failure": "WARNING",
    "link_flap":            "WARNING",
    "memory_warning":       "WARNING",
    "ospf_neighbor_down":   "CRITICAL",
    "bgp_session_drop":     "CRITICAL",
    "error_generic":        "ERROR",
}


# ── Parser ─────────────────────────────────────────────────────────────────────
def parse_log_file(filepath: Path) -> list:
    """Parse a log file and return list of detected events."""
    events = []
    device = filepath.stem

    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()
    except Exception as e:
        log.error(f"Cannot read {filepath}: {e}")
        return events

    for lineno, line in enumerate(lines, 1):
        line = line.strip()
        if not line:
            continue
        for event_type, pattern in PATTERNS.items():
            m = pattern.search(line)
            if m:
                groups = m.groupdict()
                events.append({
                    "device":     groups.get("device", device),
                    "event_type": event_type,
                    "severity":   SEVERITY.get(event_type, "INFO"),
                    "timestamp":  groups.get("timestamp", ""),
                    "interface":  groups.get("interface", ""),
                    "value":      groups.get("value", ""),
                    "user":       groups.get("user", ""),
                    "raw_line":   line[:200],
                    "file":       filepath.name,
                    "line_no":    lineno,
                })
                break  # one match per line

    log.info(f"Parsed {filepath.name}: {len(events)} events found.")
    return events


# ── Anomaly Detection ──────────────────────────────────────────────────────────
def detect_anomalies(events: list) -> list:
    """Flag anomalies based on thresholds from config."""
    thresholds = CONFIG.get("thresholds", {})
    anomalies  = []

    # Count events per type
    counts = defaultdict(int)
    for e in events:
        counts[e["event_type"]] += 1

    for event_type, count in counts.items():
        threshold = thresholds.get(event_type, thresholds.get("default", 5))
        if count >= threshold:
            anomalies.append({
                "event_type": event_type,
                "count":      count,
                "threshold":  threshold,
                "severity":   SEVERITY.get(event_type, "WARNING"),
                "message":    f"{event_type.replace('_',' ').title()} occurred {count}x (threshold: {threshold})"
            })
            log.warning(f"ANOMALY: {event_type} — {count} occurrences (threshold: {threshold})")

    # Check CPU/memory values
    for e in events:
        if e["event_type"] in ("high_cpu", "memory_warning") and e.get("value"):
            try:
                val = int(e["value"])
                limit = thresholds.get(e["event_type"] + "_pct", 85)
                if val >= limit:
                    anomalies.append({
                        "event_type": e["event_type"],
                        "count":      1,
                        "threshold":  limit,
                        "severity":   "CRITICAL" if val >= 95 else "WARNING",
                        "message":    f"{e['event_type'].replace('_',' ').title()}: {val}% on {e['device']} (limit: {limit}%)"
                    })
            except ValueError:
                pass

    return anomalies


# ── Excel Export ───────────────────────────────────────────────────────────────
SEV_COLORS = {
    "CRITICAL": "3B0D0D",
    "WARNING":  "3B2A00",
    "ERROR":    "1A1A3B",
    "INFO":     "141414",
}
HEADER_FILL = PatternFill("solid", fgColor="FF6B35")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

def export_excel(events: list, anomalies: list) -> Path:
    wb = Workbook()

    # Events sheet
    ws = wb.active
    ws.title = "Events"
    ws.sheet_view.showGridLines = False

    cols = ["device", "event_type", "severity", "timestamp", "interface", "value", "file", "line_no", "raw_line"]
    for ci, col in enumerate(cols, 1):
        c = ws.cell(1, ci, col.replace("_", " ").title())
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(len(col) + 4, 14)

    for ri, ev in enumerate(events, 2):
        fill = PatternFill("solid", fgColor=SEV_COLORS.get(ev["severity"], "141414"))
        for ci, col in enumerate(cols, 1):
            c = ws.cell(ri, ci, str(ev.get(col, "")))
            c.font = Font(color="F0EEE8", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="center")

    # Anomalies sheet
    ws2 = wb.create_sheet("Anomalies")
    ws2.sheet_view.showGridLines = False
    a_cols = ["event_type", "count", "threshold", "severity", "message"]
    for ci, col in enumerate(a_cols, 1):
        c = ws2.cell(1, ci, col.replace("_", " ").title())
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        ws2.column_dimensions[get_column_letter(ci)].width = 22

    for ri, an in enumerate(anomalies, 2):
        fill = PatternFill("solid", fgColor=SEV_COLORS.get(an["severity"], "141414"))
        for ci, col in enumerate(a_cols, 1):
            c = ws2.cell(ri, ci, str(an.get(col, "")))
            c.font = Font(color="F0EEE8", size=10)
            c.fill = fill

    path = OUTPUT_DIR / f"log_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    log.info(f"Excel report saved: {path.name}")
    return path


# ── Email Alert ────────────────────────────────────────────────────────────────
def send_alert(anomalies: list):
    cfg = CONFIG.get("email", {})
    if not cfg.get("enabled", False) or not anomalies:
        return

    critical = [a for a in anomalies if a["severity"] == "CRITICAL"]
    subject  = f"[ALERT] {len(anomalies)} anomalies detected ({len(critical)} CRITICAL)"

    body = f"Log Analyzer Alert — {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
    body += f"Total anomalies: {len(anomalies)}\n"
    body += f"Critical: {len(critical)}\n\n"
    body += "─" * 50 + "\n"
    for a in anomalies:
        body += f"[{a['severity']}] {a['message']}\n"

    msg = MIMEMultipart()
    msg["From"]    = cfg["sender"]
    msg["To"]      = ", ".join(cfg["recipients"])
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    try:
        with smtplib.SMTP(cfg["smtp_host"], cfg.get("smtp_port", 587)) as server:
            server.starttls()
            server.login(cfg["sender"], cfg["password"])
            server.sendmail(cfg["sender"], cfg["recipients"], msg.as_string())
        log.info(f"Alert email sent to: {cfg['recipients']}")
    except Exception as e:
        log.error(f"Email failed: {e}")


# ── CLI ────────────────────────────────────────────────────────────────────────
def build_cli():
    parser = argparse.ArgumentParser(
        description="Log Analyzer — parse network device logs and detect anomalies"
    )
    parser.add_argument("logs", nargs="*", default=[], help="Log files or directories to analyze")
    parser.add_argument("--dir",    "-d", default="logs", help="Directory containing log files (default: logs/)")
    parser.add_argument("--output", "-o", default="output", help="Output directory (default: output/)")
    parser.add_argument("--no-email", action="store_true", help="Skip email alerts")
    parser.add_argument("--summary", "-s", action="store_true", help="Print summary only")
    return parser


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    parser = build_cli()
    args   = parser.parse_args()

    log.info("=== Log Analyzer started ===")

    # Collect log files
    log_files = []
    if args.logs:
        for item in args.logs:
            p = Path(item)
            if p.is_dir():
                log_files.extend(p.glob("*.log"))
                log_files.extend(p.glob("*.txt"))
            elif p.exists():
                log_files.append(p)
    else:
        log_dir = Path(args.dir)
        log_dir.mkdir(exist_ok=True)
        log_files = list(log_dir.glob("*.log")) + list(log_dir.glob("*.txt"))

    if not log_files:
        log.warning(f"No log files found. Add .log files to '{args.dir}/' or pass them as arguments.")
        log.info("Tip: run  py generate_sample_logs.py  to create test logs.")
        return

    # Parse
    all_events = []
    for lf in log_files:
        all_events.extend(parse_log_file(lf))

    log.info(f"Total events parsed: {len(all_events)}")

    # Detect anomalies
    anomalies = detect_anomalies(all_events)

    # Summary
    counts = defaultdict(int)
    for e in all_events:
        counts[e["severity"]] += 1

    print("\n" + "═" * 50)
    print("  LOG ANALYZER — SUMMARY")
    print("═" * 50)
    print(f"  Log files analyzed : {len(log_files)}")
    print(f"  Total events       : {len(all_events)}")
    print(f"  CRITICAL           : {counts['CRITICAL']}")
    print(f"  WARNING            : {counts['WARNING']}")
    print(f"  ERROR              : {counts['ERROR']}")
    print(f"  Anomalies detected : {len(anomalies)}")
    print("═" * 50 + "\n")

    if anomalies:
        print("ANOMALIES:")
        for a in anomalies:
            print(f"  [{a['severity']}] {a['message']}")
        print()

    if not args.summary:
        export_excel(all_events, anomalies)

    if not args.no_email:
        send_alert(anomalies)

    log.info("=== Log Analyzer completed ===")


if __name__ == "__main__":
    main()
